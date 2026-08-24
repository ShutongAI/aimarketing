# -*- coding: utf-8 -*-
"""盒马门店业态归属预测 规则引擎 v3
v3 修复: P0 括号剥离过度触发 —— 括号内若是"纯业态词"或"仓储designation"，必须并入主干判定；
         若是"地名中夹带业态词"，则不参与判定但标记业态词冲突需复核。
X列仅4值: 盒马鲜生 / 盒马mini / 超盒算NB / 独立（非盒马系列）
"""
import re, sys, collections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

XS, MINI, NB, DL = "盒马鲜生", "盒马mini", "超盒算NB", "独立（非盒马系列）"

CITY  = r'^(上海|杭州|南京|宁波|绍兴|湖州|嘉兴|无锡|苏州|台州|宿迁|深圳|佛山|南通|北京|武汉|绵阳|桐乡|海门)市?'
MALL  = r'^(大都荟|季佳荟|馥邦|香悦里)'
NOISE = r'(急招|新店开业|招聘|开业|新店)'
WARE  = r'(前置仓|配送仓|云仓|仓库|配送站|配送点|线下配送点|配送专用|加工中心|分拨中心|分拨)'
DEAD  = r'(盒马邻里|盒马集市|盒马菜市|盒马来了|盒马鲜厨|盒马小站|盒马f2)'
# 地名护栏：这些是商场名/地名形态，不是业态词
LOCG  = r'(邻里中心|邻里商业|邻里店|梧桐邻里|菜市店|菜市场|折扣广场|城市奥莱|悦尚奥莱|奥莱曹路|奥莱万祥|奥莱星悦城)'
# 可被识别的业态token（用于括号内容分类）
FMT   = r'(会员店|mini|迷你|奥莱折扣|奥莱|奥菜|奧莱|超盒算nb|超盒算|超合算|盒马nb|nb|premier|黑标|配送专用|前置仓|配送仓|云仓|仓库|配送站|配送点|加工中心)'
# 主干是否已有可判别的业态信号
DISCRIM = r'(nb|超盒算|超合算|mini|迷你|会员店|奥莱|奥菜|奧莱|premier|黑标|盒马鲜生|盒马生鲜|盒马超市|平价社区超市|盒马邻里|盒马集市|盒马菜市|盒马来了|盒马鲜厨)' + '|' + WARE


def classify_paren(p):
    """括号内容分类 -> ('format'|'warehouse'|'ambiguous'|'location', 词)"""
    pl = re.sub(r'[\s\-_·、,，]', '', p.lower().strip())
    core = re.sub(r'(第\d+)?店$', '', pl).strip()
    # 先整体比对，再比对去"店"后缀的形态：避免把"会员店"的"店"误当门店后缀吃掉
    if re.fullmatch(FMT, pl):                         # A类: 括号内就是纯业态词
        return ('format', pl)
    if re.fullmatch(FMT, core):
        return ('format', core)
    m = re.search(r'(前置仓|配送仓|云仓|配送站|配送点|加工中心)', pl)
    if m:                                             # 仓储designation夹在地名里也成立
        return ('warehouse', m.group(1))
    if re.search(LOCG, pl):                           # 地名护栏优先
        return ('location', None)
    m = re.search(FMT, pl)
    if m:                                             # B类: 地名中夹带业态词 -> 歧义
        return ('ambiguous', m.group(1))
    return ('location', None)


def preclean(name, cleaned):
    raw = str(name or '') or str(cleaned or '')
    t = raw.replace('（', '(').replace('）', ')').lower()
    parens = re.findall(r'\(([^)]*)\)', t)
    trunk = re.sub(r'\([^)]*\)', '', t)
    trunk = re.sub(NOISE, '', trunk)
    trunk = re.sub(CITY, '', trunk)
    trunk = re.sub(MALL, '', trunk)
    trunk = re.sub(r'[\s\-_·、,，]', '', trunk).strip()

    merged, amb = trunk, ''
    has_sig = bool(re.search(DISCRIM, trunk))
    for p in parens:
        kind, w = classify_paren(p)
        if kind in ('format', 'warehouse'):
            merged += w                                # 并入主干
        elif kind == 'ambiguous':
            if not has_sig:
                merged += w                            # 主干无信号时，采纳括号内业态词
            else:
                amb = w                                # 主干已有信号 -> 记为冲突，待复核
    return merged, amb


def predict(name, cleaned):
    b, amb = preclean(name, cleaned)
    c = re.sub(r'[\s\-_·、,，]', '', str(cleaned or '').lower())
    both = (b if b else c) + '|' + c

    def out(x, y, z, cf):
        if amb:
            cf = '低'
            z += f"；注意：名称内另含业态词「{amb}」，存在冲突"
        return x, y, z, cf

    # PA 非门店实体：仓储配送
    m = re.search(WARE, both)
    if m:
        return out(DL, m.group(1), f"命中仓储配送词「{m.group(1)}」，盒马体系内非零售实体，从门店盘剔除", "高")
    # PA2 第三方服务（非盒马官方门店）
    if '代购' in both:
        return out(DL, "代购", "命中「代购」，第三方代购服务点，非盒马官方门店", "低")

    # P1 超盒算NB 正向硬命中
    if re.search(r'超盒算', both):  return out(NB, "", "命中「超盒算」官方现用名", "高")
    if re.search(r'超合算', both):  return out(NB, "", "命中「超合算」(超盒算采集错别字)", "高")
    if re.search(r'盒马nb', both):  return out(NB, "", "命中「盒马NB」，NB业态并称", "高")
    if re.search(r'(?<![a-z])nb(?![a-z])', both):
        return out(NB, "", "命中光杆「NB」独立token，本样本光杆NB全部为超盒算NB门店", "高")
    if re.search(r'奥莱折扣', both):
        return out(NB, "", "命中「奥莱折扣」，奥莱折扣店即超盒算NB前身，同店同址", "高")
    if re.search(r'平价社区超市', both):
        return out(NB, "", "命中「平价社区超市」，超盒算NB官方定位话术(内证:样本存在NB旗下平价社区超市)", "中")

    # P2 奥莱系 → 更名为超盒算NB
    m = re.search(r'(生鲜奥莱|盒马奥莱|奥莱自提|生鲜奥菜|生鲜奧莱|奥菜|奧莱|奥莱|奥特莱斯)', both)
    if m and not (re.search(LOCG, both) and not re.search(r'(生鲜奥莱|盒马奥莱|奥莱折扣|奥菜|奧莱)', both)):
        return out(NB, "奥莱", f"命中已淘汰业态词「{m.group(1)}」，盒马生鲜奥莱已整体更名为超盒算NB", "低")

    # P3 盒马mini
    if re.search(r'mini|迷你', both):
        return out(MINI, "", "命中「mini」小业态", "高")

    # P4 会员店 / Premier → 转制为盒马鲜生
    if re.search(r'会员店', both):
        return out(XS, "会员店", "命中已淘汰业态词「会员店」，盒马X会员店已淘汰，存量门店转制为盒马鲜生", "低")
    m = re.search(r'(premier|黑标)', both)
    if m:
        return out(XS, "Premier黑标店", f"命中已淘汰业态词「{m.group(1)}」，高端精品业态已淘汰，转制为盒马鲜生", "低")

    # P6 已关停子业态（整词匹配）
    m = re.search(DEAD, both)
    if m:
        return out(DL, m.group(1).replace('盒马', ''), f"命中已关停子业态「{m.group(1)}」，社区团购/非商超实体，非在营门店", "中")

    # P7 盒马鲜生 主力业态
    if re.search(r'盒马鲜生|盒马生鲜|盒马超市|盒马旗下', both):
        return out(XS, "", "命中「盒马鲜生/盒马超市」主力大店业态", "高")
    if re.search(r'^盒马$|^盒马\(', b) or c == '盒马':
        return out(XS, "", "光杆「盒马」无其它业态特征词，按主力业态盒马鲜生兜底", "低")
    if '盒马' in both:
        return out(XS, "", "含盒马字样但无明确业态特征词，按主力业态盒马鲜生兜底", "低")

    # P8 无盒马标识
    return out(DL, "无盒马字样", "名称中无盒马/超盒算标识，判为非盒马系列", "中")


def main(src, dst):
    wb = openpyxl.load_workbook(src); ws = wb['Sheet1']
    pc = collections.Counter(ws.cell(r, 23).value for r in range(2, ws.max_row + 1))
    GENERIC = {u for u, n in pc.items() if u and n >= 10}

    for col, h in ((24,'连锁名称预测(X)'),(25,'负向词/黑名单(Y)'),(26,'判断依据(Z)'),(27,'置信度'),(28,'复核建议')):
        ws.cell(1, col).value = h

    stat=collections.Counter(); yst=collections.Counter(); conf=collections.Counter(); review=[]
    for r in range(2, ws.max_row + 1):
        g = lambda c: ws.cell(r, c).value
        pic = g(23); has_pic = bool(pic) and pic not in GENERIC
        x, y, z, cf = predict(g(3), g(22))
        if cf == '高' and not has_pic: cf = '中'
        need = ''
        if y in ('奥莱','会员店','Premier黑标店'): need = '需图片复核(已淘汰业态,确认转制后招牌)'
        elif cf == '低': need = '建议人工复核'
        elif not has_pic and '兜底' in z: need = '建议人工复核'
        for col, v in ((24,x),(25,y),(26,z),(27,cf),(28,need)): ws.cell(r, col).value = v
        stat[x]+=1; conf[cf]+=1
        if y: yst[y]+=1
        if need: review.append([r, g(3), g(10), x, y, need, pic or '', '有效图' if has_pic else '无有效图'])

    rs = wb.create_sheet('需复核清单')
    rs.append(['原表行号','店铺名称','城市','规则判定X','Y负向词','复核类型','门头照URL','图片可用性'])
    for row in review: rs.append(row)
    for i,w in enumerate([10,36,12,16,14,32,58,12],1):
        rs.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    for c in rs[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='C00000')
    rs.freeze_panes='A2'

    ss = wb.create_sheet('预测结果汇总'); N = ws.max_row-1
    ss.append(['X 连锁名称','门店数','占比'])
    for k,v in stat.most_common(): ss.append([k,v,f'{v/N*100:.1f}%'])
    ss.append([]); ss.append(['Y 负向词','命中数','占比'])
    for k,v in yst.most_common(): ss.append([k,v,f'{v/N*100:.1f}%'])
    ss.append([]); ss.append(['置信度','行数','占比'])
    for k in ('高','中','低'): ss.append([k,conf[k],f'{conf[k]/N*100:.1f}%'])
    ss.append([]); ss.append(['需复核合计',len(review),f'{len(review)/N*100:.1f}%'])
    for i,w in enumerate([26,12,10],1): ss.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    for c in ss[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='2F5597')

    for c in ws[1]: c.font=Font(bold=True); c.alignment=Alignment('center','center',wrap_text=True)
    for i,w in ((24,18),(25,16),(26,60),(27,9),(28,32)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    ws.freeze_panes='A2'; wb.save(dst)

    print(f'总行数 {N}')
    print('\n=== X 列分布 ===')
    for k,v in stat.most_common(): print(f'{k:<20}{v:>6}  {v/N*100:>5.1f}%')
    print('\n=== Y 列负向词 ===')
    for k,v in yst.most_common(): print(f'{k:<18}{v:>6}')
    print('\n=== 置信度 ===')
    for k in ('高','中','低'): print(f'{k:<20}{conf[k]:>6}  {conf[k]/N*100:>5.1f}%')
    print(f'\n需复核 {len(review)} 行')
    assert set(stat) <= {XS,MINI,NB,DL}; print('✓ X列封闭性校验通过')

main(sys.argv[1], sys.argv[2])

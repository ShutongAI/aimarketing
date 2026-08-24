# -*- coding: utf-8 -*-
import openpyxl, re, collections
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ws = openpyxl.load_workbook('盒马门店业态归属预测_全量结果.xlsx')['Sheet1']
rows = [(str(ws.cell(r,3).value or ''), str(ws.cell(r,22).value or ''), ws.cell(r,24).value)
        for r in range(2, ws.max_row+1)]
BR = ['盒马鲜生','盒马mini','超盒算NB','独立（非盒马系列）']

def stat(w):
    rx = re.compile(r'(?<![a-z])nb(?![a-z])' if w.lower()=='nb' else re.escape(w.lower()))
    hit = [x for n,c,x in rows if rx.search((n+'|'+c).lower())]
    cc = collections.Counter(hit)
    return len(hit), cc

# ---- 正向词：品牌 -> [词]
POS = {
 '盒马鲜生': ['盒马鲜生','盒马超市','会员店','盒马鲜生会员店','X会员店','Premier','黑标'],
 '盒马mini': ['mini','迷你','盒马mini','盒马鲜生mini','盒马生鲜mini'],
 '超盒算NB': ['超盒算','超盒算NB','盒马NB','NB(光杆)','超合算','奥莱','生鲜奥莱','盒马奥莱','盒马生鲜奥莱',
              '奥莱折扣','奥菜','奧莱','平价社区超市','盒马旗下平价社区超市','盒马旗下'],
}
# ---- 负向词：品牌 -> [词]（含该词则大概率不是本品牌）
NEG = {
 '盒马鲜生': ['NB','超盒算','超合算','盒马NB','奥莱','生鲜奥莱','奥莱折扣','奥菜','奧莱','平价社区超市','盒马旗下',
              'mini','迷你','前置仓','配送仓','云仓','仓库','配送站','配送点','配送专用','加工中心','代购',
              '盒马邻里','盒马集市','盒马菜市','盒马来了','盒马鲜厨'],
 '盒马mini': ['NB','超盒算','超合算','盒马NB','奥莱','生鲜奥莱','奥莱折扣','平价社区超市','盒马旗下','会员店',
              'Premier','黑标','前置仓','配送仓','云仓','仓库','配送站','配送点','配送专用','加工中心','代购',
              '盒马邻里','盒马集市','盒马菜市','盒马来了','盒马鲜厨'],
 '超盒算NB': ['会员店','盒马鲜生会员店','mini','迷你','Premier','黑标','盒马超市',
              '前置仓','配送仓','云仓','仓库','配送站','配送点','配送专用','加工中心','代购',
              '盒马邻里','盒马集市','盒马菜市','盒马来了','盒马鲜厨'],
}
# ---- 待定（吃不准，先不定义）
PEND = [
 ('超市','无区分度','70行: 盒马鲜生33 / 超盒算NB37 (52.9%)','需与"盒马超市""平价社区超市"连用才有判别力'),
 ('生鲜','被奥莱污染','56行: NB43 / 鲜生12 (76.8%)','"生鲜奥莱"拉走了大部分命中，单用不可靠'),
 ('盒马生鲜','条件成立','53行: NB42 / 鲜生10 (裸词纯度仅18.9%)；剔除"奥莱"后 鲜生10/11=91%','已从盒马鲜生正向词中移出。仅在不含"奥莱"时才可用作盒马鲜生正向词'),
 ('社区','来源单一','28行: NB26 (92.9%)','判别力全部来自"平价社区超市"；裸"社区"多为小区名(如香醍国际社区七期)'),
 ('奥特莱斯','样本过小+地名嫌疑','3行: 鲜生2 / NB1 (66.7%)','多为"XX奥特莱斯店"商场名，非业态词'),
 ('折扣','样本小','6行: NB6 (100%)','仅"奥莱折扣"连用可信；"XX折扣广场店"是商场名'),
 ('邻里','地名护栏词','17行: NB16 / 独立1','仅"盒马邻里"整词是业态词；"XX邻里中心/邻里店"是商场名，误伤率88%'),
 ('菜市','地名护栏词','4行: 独立3 / NB1','仅"盒马菜市"整词是业态词；"XX菜市店"是地名'),
 ('生活','无区分度','12行: 鲜生6 / NB5 / mini1 (50%)','多为"XX生活广场"商场名'),
 ('自提','样本过小','1行: NB1','仅"奥莱自提"1例，不足以定义'),
 ('滴答生鲜超市','疑似第三方','1行: 盒马滴答生鲜超市','疑似蹭牌店或联营，需看门头照定性'),
 ('略逮禾生鲜超市','疑似第三方','1行: 略逮盒马禾生鲜超市','名称异常，疑似采集错误或第三方，需人工核'),
 ('大都荟/季佳荟/馥邦/香悦里','商场名前缀','4行','商场名+盒马，主词仍是盒马但无业态词，需看门头照定业态'),
 ('园区/站/点(单字)','地名嫌疑','—','"汉南园区""上马站"是地名；仅"配送站/配送点"整词算仓储'),
 ('盒马(光杆)','无业态信号','64行 全部兜底判盒马鲜生','已从盒马鲜生正向词中移出。按主力业态兜底，但无法排除实为NB/mini，建议看门头照'),
]

wb = openpyxl.Workbook()
H = Font(bold=True, color='FFFFFF'); HF = PatternFill('solid', fgColor='2F5597')
BD = Border(*[Side('thin', color='D0D0D0')]*4)
def fin(ws_, widths, hcolor='2F5597'):
    for i,w in enumerate(widths,1):
        ws_.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for c in ws_[1]:
        c.font=H; c.fill=PatternFill('solid',fgColor=hcolor)
        c.alignment=Alignment('center','center',wrap_text=True)
    for row in ws_.iter_rows(min_row=2):
        for c in row: c.border=BD; c.alignment=Alignment(vertical='top', wrap_text=True)
    ws_.freeze_panes='A2'

# ===== Sheet1 主表：三列 =====
s = wb.active; s.title='盒马三品牌正负向词'
s.append(['品牌名','正向词（poi_name 含则大概率是本品牌）','负向词（poi_name 含则大概率不是本品牌）'])
for b in ['盒马鲜生','盒马mini','超盒算NB']:
    s.append([b, '、'.join(POS[b]), '、'.join(NEG[b])])
fin(s,[14,52,68])
for r in range(2,5): s.row_dimensions[r].height=96

# ===== Sheet2 正向词明细 =====
s = wb.create_sheet('正向词明细')
s.append(['品牌名','正向词','本样本命中行数','本品牌占比(纯度)','鲜生','mini','NB','独立','备注'])
for b in ['盒马鲜生','盒马mini','超盒算NB']:
    for w in POS[b]:
        probe = {'盒马(光杆)':'盒马','NB(光杆)':'nb','X会员店':'会员店'}.get(w, w)
        n, cc = stat(probe)
        pur = f'{cc[b]/n*100:.1f}%' if n else '—'
        note = ''
        if w=='盒马(光杆)': note='需为独立词"盒马"，含其它业态词时让位'
        if w=='NB(光杆)': note='独立token匹配，前后不可接字母'
        if w=='盒马生鲜': note='⚠ 仅在不含"奥莱"时成立（否则指向NB）'
        if w in ('奥莱','生鲜奥莱','盒马奥莱','盒马生鲜奥莱','奥菜','奧莱'): note='已淘汰业态→更名NB，建议看门头照确认'
        if w in ('会员店','盒马鲜生会员店','X会员店','Premier','黑标'): note='已淘汰业态→转制鲜生，建议看门头照确认'
        s.append([b,w,n,pur,cc[BR[0]],cc[BR[1]],cc[BR[2]],cc[BR[3]],note])
fin(s,[12,22,13,14,8,8,8,8,40],'375623')

# ===== Sheet3 负向词明细 =====
s = wb.create_sheet('负向词明细')
s.append(['品牌名','负向词','本样本命中行数','命中行中属本品牌的行数','排除力','该词实际指向','例外说明'])
for b in ['盒马鲜生','盒马mini','超盒算NB']:
    for w in NEG[b]:
        probe = {'NB':'nb'}.get(w, w)
        n, cc = stat(probe)
        own = cc[b]
        power = '强(100%排除)' if own==0 else f'中(仍有{own}行属本品牌)'
        top = cc.most_common(1)[0][0] if n else '—'
        exc = ''
        if b=='盒马鲜生' and w=='mini':
            exc = '唯一例外: 盒马鲜生(上海江桥恒久mini店) —— mini在括号内地名中，已标记需复核'
        s.append([b,w,n,own,power,top,exc])
fin(s,[12,22,13,20,18,20,46],'843C0C')

# ===== Sheet4 待定 =====
s = wb.create_sheet('待定_需人工复核')
s.append(['待定词','待定原因','本样本实测分布','说明（暂不定义为正/负向词）'])
for r in PEND: s.append(list(r))
fin(s,[24,18,46,54],'7030A0')

# ===== Sheet5 使用规则 =====
s = wb.create_sheet('使用规则')
s.append(['序号','规则','说明/证据'])
for r in [
 (1,'负向词优先级 > 正向词','同时命中时以负向词为准。证据：盒马鲜生NB店→NB、盒马鲜生奥莱→NB、盒马鲜生mini→mini，6行交叉全部由负向词正确改判'),
 (2,'负向词判定顺序：NB > 奥莱 > mini > 会员店','大量NB门店名里仍挂"奥莱折扣"旧词，NB不排在最前会误判'),
 (3,'仓储配送词优先级最高','前置仓/配送仓/云仓/仓库/配送站/配送点/配送专用/加工中心——任何业态的仓都不是门店，先剔除再谈业态'),
 (4,'"盒马鲜生"不是NB/mini的负向词','含"盒马鲜生"的行里有4行属NB、2行属mini，它只是较弱的正向词，会让位于更强信号'),
 (5,'地名护栏：先定作用域再匹配','邻里/菜市/折扣/奥莱/园区/站 在"XX中心店/XX广场店"位置上是商场名，不是业态词。裸匹配"邻里"误伤率88%'),
 (6,'括号内容分三类处理','A纯业态词(会员店)并入判定；B仓储标注并入；C地名夹带业态词不参与判定。一律剥离会漏判13行盒马鲜生(会员店)'),
 (7,'V列(poi_name_cleaned)可直接用','V列已做过正则清洗，正向词在V列上匹配纯度更高；但城市前缀(上海/杭州…)仍需剥离'),
 (8,'错别字/异体字必须覆盖','超合算=超盒算、奥菜/奧莱=奥莱、Nb/nb=NB'),
]: s.append(list(r))
fin(s,[6,40,86],'C00000')

wb.save('盒马三品牌正负向词表_v1.xlsx')
print('saved')
for n in wb.sheetnames: print(' -', n, wb[n].max_row-1, '行')
